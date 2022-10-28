import datetime
import openpyxl


class Currency:
  def __init__(self, curr_name, df_complete, df_refund):
    self.curr_name = curr_name
    self.df_turnover = df_complete
    self.df_turnover_refund = df_refund
    self.EU = ['AUSTRIA', 'BELGIUM', 'BULGARIA', 'CROATIA', 'CYPRUS', 'CZECH REPUBLIC', 'DENMARK', 'ESTONIA', 'FINLAND',
          'FRANCE', 'GERMANY', 'GREECE', 'HUNGARY', 'IRELAND', 'ITALY', 'LATVIA', 'LITHUANIA', 'LUXEMBOURG', 'MALTA',
          'NETHERLANDS', 'POLAND', 'ROMANIA', 'PORTUGAL', 'SPAIN', 'SLOVAKIA', 'SLOVENIA', 'SWEDEN']
    self.cup = ['CHINA UNION PAY']

#  complete turnover by currency
  def turnover_comp(self):
    df_turnover_comp = self.df_turnover[self.df_turnover['Валюта'].isin(self.curr_name)]
    turnover_curr = df_turnover_comp['Стоимость операции'].sum()
    return turnover_curr

# refund turnover by currency
  def turnover_refund(self):
    df_turnover_ref = self.df_turnover_refund[self.df_turnover_refund['Валюта'].isin(self.curr_name)]
    turnover_ref_curr = df_turnover_ref['Стоимость операции'].sum()  # оборот рефандов Евро
    return turnover_ref_curr


  def turnover_comp_EU(self):  # Turnover for EU countries
    df_turnover_comp = self.df_turnover[(self.df_turnover['Валюта'].isin(self.curr_name)) & (self.df_turnover['Название страны-эмитента'].isin(self.EU))]
    turnover_curr_eu = df_turnover_comp['Стоимость операции'].sum()
    return turnover_curr_eu

  def turnover_comp_NonEU(self):
    df_turnover_comp = self.df_turnover[(self.df_turnover['Валюта'].isin(self.curr_name)) & (~self.df_turnover['Название страны-эмитента'].isin(self.EU))]
    turnover_curr_non_eu = df_turnover_comp['Стоимость операции'].sum()
    return turnover_curr_non_eu

  def turnover_refund_EU(self):
    df_turnover_ref = self.df_turnover_refund[(self.df_turnover_refund['Валюта'].isin(self.curr_name)) & (self.df_turnover_refund['Название страны-эмитента'].isin(self.EU))]
    turnover_ref_curr_EU = df_turnover_ref['Стоимость операции'].sum()  # оборот рефандов Евро
    return turnover_ref_curr_EU

  def turnover_refund_NonEU(self):
    df_turnover_ref = self.df_turnover_refund[(self.df_turnover_refund['Валюта'].isin(self.curr_name)) & (~
      self.df_turnover_refund['Название страны-эмитента'].isin(self.EU))]
    turnover_ref_curr_NonEU = df_turnover_ref['Стоимость операции'].sum()  # оборот рефандов Евро
    return turnover_ref_curr_NonEU

#  refund count by currency

  def count_refund(self):
    #  ниже датафрейм из рефандов евро и без КУП

    refund_count_df = self.df_turnover_refund[self.df_turnover_refund['Валюта'].isin(self.curr_name)]
    refund_count_curr = refund_count_df['Стоимость операции'].count()  # кол-во рефандов by currency
    return refund_count_curr

  # Quantity of transactions
  def count_complete(self):
    return self.df_turnover['Стоимость операции'].count()


def fee(df_turnover, rate):
    return df_turnover * rate


def hold(df_turnover, fee, hrate):
    return (df_turnover - fee) * hrate


def wb_holdback(statement_sheet, date_from_obj, date_to_obj, now_date,time):
  curr_names = []
  curr_values = []

  date_hold = date_to_obj + datetime.timedelta(time)
  for i in range(73, 75):                                         # hold size in statement
    curr_name = statement_sheet.cell(row=i, column=1).value
    curr_value = statement_sheet.cell(row=i, column=3).value
    curr_names.append(curr_name)
    curr_values.append(curr_value)
  hold_dict = dict(zip(curr_names, curr_values))
  print(hold_dict)
  hold = 'Hold.xlsx'  # Name of Excel file (rename)                                           #REPLACE MERCH NAME
  hbook = openpyxl.load_workbook(filename=hold)
  hold_sheet = hbook['WB']
  for i in range(5, 1000):  # search an empty raw
    value = hold_sheet.cell(row=i, column=1).value
    if value is None:
      hold_sheet.cell(row=i, column=1).value = 'Whitebit'
      hold_sheet.cell(row=i, column=2).value = 'Period from ' + str(date_from_obj.date()) + ' until ' + str(date_to_obj.date())
      hold_sheet.cell(row=i, column=8).value = date_hold
      break
  for curr in hold_dict:
    for j in range(11, 14):                             # currency size in hold file
      value = hold_sheet.cell(row=2, column=j).value  # search a currency
      if value == curr:
        hold_sheet.cell(row=i, column=j).value = hold_dict.get(curr)  # write a hold value

  # holback return
  curr_names = []
  curr_values = []
  holdreturn_dict = {}


  # Using 0 in case holdbackreturn = none                              EXPAND FOR OTHER MERCHANT
  statement_sheet['C81'] = 0
  statement_sheet['C82'] = 0

  for m in range(4, 30):  # search the holdback rerturn date
    value = hold_sheet.cell(row=m, column=8).value
    value1 = hold_sheet.cell(row=m+1, column=8).value

    if value1 > date_to_obj and value <= date_to_obj:
      hold_sheet.cell(row=m, column=9).value = str(date_to_obj.date())      # write a holdback return date
      hold_sheet.cell(row=m, column=10).value = now_date # write a current date
      for n in range(11, 42):                            # currency size in hold file
        curr_name = hold_sheet.cell(row=2, column=n).value
        curr_value = hold_sheet.cell(row=m, column=n).value

        curr_names.append(curr_name)
        curr_values.append(curr_value)
      holdreturn_dict = dict(zip(curr_names, curr_values))  # dict with holdback returns
      print(holdreturn_dict)
      break

  for curr in holdreturn_dict:
    for i in range(81, 85):           #holdback return range in Excel
      value = statement_sheet.cell(row=i, column=1).value  # search a currency
      if value == curr:
        statement_sheet.cell(row=i, column=3).value = holdreturn_dict.get(curr)  # write a hold value

  hbook.save(hold)
